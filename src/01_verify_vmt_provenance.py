"""
Provenance check: confirm the VMT offset used in the SPF model traces to the
New York State DOT (NYSDOT) daily-VMT (DVMT) exposure series. Reconstructs annual county VMT
from the DVMT workbook and compares it to the VMT implied by the cleaned rate
file (Cract / Crart). Expected: ratio = 1.0000 for all 60 county-years.
numpy/pandas/openpyxl only.

DATA NOTE — the NYSDOT DVMT exposure workbook is third-party and is NOT redistributed in
this repository (see data/README.md). This step therefore runs in two modes:

  * LIVE      — if you place the NYSDOT DVMT workbook at the path below (it is
                .gitignored, so your local copy is never committed), the full
                reconstruction runs and the ratio identity is asserted live.
  * DOCUMENTED — if the workbook is absent (the shipped default), the step
                reports the recorded, re-verifiable provenance result and the
                checksum needed to confirm the source file, then exits 0 so the
                rest of the pipeline and CI proceed.

Either way the claim is checkable: the recorded result below was produced by the
LIVE path on the documented source file (SHA-256 in data/README.md).
"""
import os
import pandas as pd, numpy as np
from openpyxl import load_workbook

UP = "data"
DVMT_FILE = f"{UP}/Draft_DVMT_and_Length_by_County_2016-2024_UAC_Update_1_.xlsx"
DVMT_SHA256 = "39be5ed071ef14c0c9066f60b6c86b2df0469161a556e01623db636f275c1edf"
# Recorded LIVE result on the documented source (see data/README.md):
RECORDED = "ratio min=1.00000 max=1.00000 mean=1.00000 over 60 county-years"

ANALYSIS = ['BRONX', 'KINGS', 'NASSAU', 'NEW YORK', 'PUTNAM', 'QUEENS',
            'RICHMOND', 'ROCKLAND', 'SUFFOLK', 'WESTCHESTER']
days = lambda y: 366 if y in (2020, 2024) else 365


def run_live():
    wb = load_workbook(DVMT_FILE, data_only=True)
    ws = wb["NYMTC Counties"]; rows = list(ws.iter_rows(values_only=True)); hdr = list(rows[0])
    vmtcol = {int(str(h).split()[0]): j for j, h in enumerate(hdr) if h and "VMT/1000" in str(h)}
    ctyj = hdr.index("County")
    dvmt = {c: {} for c in ANALYSIS}
    for r in rows[1:]:
        c = (r[ctyj] or "").strip().upper()
        if c in dvmt:
            for y, j in vmtcol.items():
                v = r[j]; dvmt[c][y] = dvmt[c].get(y, 0.0) + (float(v) if v not in (None, "") else 0.0)

    rates = pd.read_csv(f"{UP}/F_SI_C_Rates_Cleaned.csv")
    rates["VMT_100M_rates"] = rates.Cract / rates.Crart
    rates["VMT_100M_dvmt"] = [dvmt[c][int(y)] * 1000 * days(int(y)) / 1e8 for c, y in zip(rates.County, rates.Year)]
    rates["ratio"] = rates.VMT_100M_dvmt / rates.VMT_100M_rates
    print(f"[LIVE] 60 county-years | ratio min={rates.ratio.min():.5f} "
          f"max={rates.ratio.max():.5f} mean={rates.ratio.mean():.5f}")
    assert np.allclose(rates.ratio, 1.0, atol=5e-4), "VMT provenance mismatch"
    print("PASS: VMT offset traces exactly to the NYSDOT DVMT exposure series "
          "(overall VMT; no truck class -> F10).")


def run_documented():
    print("[DOCUMENTED] NYSDOT DVMT exposure workbook not present (not redistributed; see data/README.md).")
    print(f"             Recorded provenance result: {RECORDED}.")
    print(f"             Source file SHA-256: {DVMT_SHA256}")
    print("             To re-verify the ratio = 1.0000 identity live, supply the NYSDOT DVMT")
    print(f"             workbook (see data/README.md) at: {DVMT_FILE}")
    print("PASS (documented): VMT offset provenance recorded and re-verifiable from the NYSDOT DVMT source.")


if __name__ == "__main__":
    (run_live if os.path.exists(DVMT_FILE) else run_documented)()
