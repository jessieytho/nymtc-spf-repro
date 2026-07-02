"""
R3 - "Wrong vs right" exhibit: a count-on-count fixed-effects (TWFE) specification
MANUFACTURES the retired ~3.9x truck "lethality multiplier" from data whose honest
conditional truck/non-truck fatality ratio is ~1.9x.

WRONG (count-on-count TWFE): regress fatal COUNTS on crash COUNTS with county+year
FE, separately for large-truck and non-truck crashes; read the ratio of the two
crash-count slopes as a "lethality multiplier."  -> 3.84x  (== retired ~3.9x).

RIGHT (conditional rate): ratio of fatal-crashes-per-crash, trucks vs non-trucks
-> 1.91x (stable); and still NOT exposure-normalized risk (no truck VMT, F10).

Why the WRONG number is an artifact (two independent reasons):
 (1) INFLATION. After county+year demeaning a count-on-count slope is
     Cov_within(fatal,crashes)/Var_within(crashes), NOT a rate. It attenuates the
     smooth high-volume non-truck series more than the lumpy rare-event truck
     series, inflating the true rate ratio ~2-fold (1.91x -> 3.84x).
 (2) NO STATISTICAL SUPPORT. The component slopes are individually insignificant
     (t ~ 0.7); the slope ratio is not stably estimable under naive-analytic,
     iid-cell, OR county-cluster inference (the non-truck denominator slope sits
     near zero and flips sign). The retired 3.9x is a point estimate masquerading
     as a finding - reproducible, yet not a valid estimand (reproduction != validation).

CIs/inference: numpy/scipy/pandas/openpyxl/matplotlib.
"""
import numpy as np, pandas as pd, os
from numpy.linalg import lstsq, inv
from openpyxl import load_workbook
import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
OUT="outputs"; UP="data"; os.makedirs(OUT,exist_ok=True)
RNG=np.random.default_rng(7); B=5000

def load(fn):
    ws=load_workbook(fn,data_only=False).active; rows=[]
    for r in ws.iter_rows(min_row=2,values_only=True):
        yr,tod,cty,dow,typ,tot=r[:6]
        rows.append((str(yr),(tod or "").strip(),(cty or "").strip(),(dow or "").strip(),(typ or "").strip(),0.0 if tot is None else float(tot)))
    return pd.DataFrame(rows,columns=["Year","TOD","County","DOW","Type","Total"])
def cy(df,a,b):
    L=df[(df.TOD=="")&(df.DOW!="")&(df.County!="")&(df.Type.isin(["Fatal","Personal Injury","Property Damage"]))]
    tot=L.groupby(["County","Year"]).Total.sum().rename(a); fat=L[L.Type=="Fatal"].groupby(["County","Year"]).Total.sum().rename(b)
    return pd.concat([tot,fat],axis=1).fillna(0).reset_index()
T=cy(load(f"{UP}/05_0506_Day_of_Week_Time_of_Day.xlsx"),"truck_cr","truck_fat")
G=cy(load(f"{UP}/01_0108_Day_of_Week_Time_of_Day.xlsx"),"gen_cr","gen_fat")
m=T.merge(G,on=["County","Year"]); m["nt_cr"]=m.gen_cr-m.truck_cr; m["nt_fat"]=m.gen_fat-m.truck_fat
counties=sorted(m.County.unique()); years=sorted(m.Year.unique())

def design(d,xcol,cidcol="County"):
    cids=sorted(d[cidcol].unique()); yrs=sorted(d.Year.unique())
    cols=[np.ones(len(d)),d[xcol].values]
    for c in cids[1:]: cols.append((d[cidcol]==c).astype(float).values)
    for y in yrs[1:]: cols.append((d.Year==y).astype(float).values)
    return np.column_stack(cols)
def slope(d,xcol,ycol,cidcol="County"):
    X=design(d,xcol,cidcol); b,*_=lstsq(X,d[ycol].values,rcond=None); return b[1]
def slope_se(d,xcol,ycol):
    X=design(d,xcol); y=d[ycol].values; b,*_=lstsq(X,y,rcond=None)
    r=y-X@b; n,k=X.shape; s2=r@r/(n-k); return b[1], np.sqrt(s2*inv(X.T@X)[1,1])
def honest(d): return (d.truck_fat.sum()/d.truck_cr.sum())/(d.nt_fat.sum()/d.nt_cr.sum())

sT,seT=slope_se(m,"truck_cr","truck_fat"); sN,seN=slope_se(m,"nt_cr","nt_fat")
sp=sT/sN; ho=honest(m)
seR=np.sqrt((seT/sN)**2+(sT*seN/sN**2)**2); naiveR=(sp-1.96*seR, sp+1.96*seR)
actT=m.truck_fat.sum()/m.truck_cr.sum(); actN=m.nt_fat.sum()/m.nt_cr.sum()

# county pairs-cluster bootstrap: honest CI + spurious instability
sp_b=[]; ho_b=[]; flip=0
for _ in range(B):
    idx=RNG.integers(0,len(counties),len(counties)); parts=[]
    for j,ci in enumerate(idx):
        s=m[m.County==counties[ci]].copy(); s["cid"]=f"d{j}"; parts.append(s)
    bd=pd.concat(parts,ignore_index=True)
    a=slope(bd,"truck_cr","truck_fat","cid"); b=slope(bd,"nt_cr","nt_fat","cid")
    if b<=0: flip+=1
    sp_b.append(a/b if b!=0 else np.nan); ho_b.append(honest(bd))
ho_ci=np.nanpercentile(ho_b,[2.5,97.5]); flip/=B

print("="*80)
print("R3  WRONG vs RIGHT - reconstructing the retired truck 'lethality multiplier'")
print("="*80)
print(f"\nWRONG count-on-count TWFE multiplier = {sp:.2f}x   (== retired ~3.9x)")
print(f"   component slopes individually insignificant: truck t={sT/seT:.1f}, nontruck t={sN/seN:.1f}")
print(f"   naive-analytic delta CI [{naiveR[0]:.1f}, {naiveR[1]:.1f}]; cluster bootstrap unstable "
      f"(denominator slope <=0 in {flip*100:.0f}% of resamples) -> no valid interval")
print(f"RIGHT conditional fatal-per-crash    = {ho:.2f}x   county-cluster 95% CI [{ho_ci[0]:.2f}, {ho_ci[1]:.2f}]")
print(f"\nMechanism (differential FE attenuation):")
print(f"  truck    FE slope {sT*1000:.2f}/1k vs actual {actT*1000:.2f}/1k (atten x{actT/sT:.1f})")
print(f"  nontruck FE slope {sN*1000:.2f}/1k vs actual {actN*1000:.2f}/1k (atten x{actN/sN:.1f})")
print(f"  -> nontruck attenuated more -> true {ho:.2f}x inflated to {sp:.2f}x (artifact)")

exhibit=pd.DataFrame([
 dict(Specification="Count-on-count TWFE (WRONG)",Multiplier=round(sp,2),
      Inference=f"slopes n.s. (t~0.7); no valid CI (analytic [{naiveR[0]:.0f},{naiveR[1]:.0f}]; cluster sign-flips {flip*100:.0f}%)",
      Verdict="reconstructs retired ~3.9x; inflated ~2x by FE attenuation; not a valid estimand"),
 dict(Specification="Conditional fatal-per-crash (RIGHT)",Multiplier=round(ho,2),
      Inference=f"county-cluster 95% CI [{ho_ci[0]:.2f}, {ho_ci[1]:.2f}] (stable, significant)",
      Verdict="honest severity contrast; still NOT exposure-normalized risk (F10)"),
])
exhibit.to_csv(f"{OUT}/r3_wrong_vs_right.csv",index=False)
print("\n"+exhibit.to_string(index=False))

# figure
fig,ax=plt.subplots(figsize=(8.4,3.5)); XMAX=5.0
ax.plot([ho_ci[0],ho_ci[1]],[0,0],color="#2f855a",lw=3.2,solid_capstyle="round")
ax.plot(ho,0,"o",color="#2f855a",ms=11,zorder=3); ax.text(ho,0.24,f"{ho:.1f}x",ha="center",color="#2f855a",weight="bold")
# spurious: point + off-scale unstable band with arrows
ax.annotate("",xy=(0.85,1),xytext=(XMAX*0.99,1),arrowprops=dict(arrowstyle="<->",color="#c53030",alpha=0.35,lw=2))
ax.plot(sp,1,"o",color="#c53030",ms=12,zorder=3); ax.text(sp,1.24,f"{sp:.1f}x",ha="center",color="#c53030",weight="bold")
ax.text((0.85+XMAX)/2,0.72,"no valid CI (slopes n.s.; ratio unstable under resampling)",
        ha="center",fontsize=8,color="#c53030",style="italic")
ax.axvline(1.0,color="#aaa",ls=":",lw=1)
ax.set_yticks([1,0]); ax.set_yticklabels(["Count-on-count TWFE\n(WRONG)","Conditional rate\n(RIGHT)"],fontsize=9.5)
ax.set_ylim(-0.7,1.8); ax.set_xlim(0.8,XMAX)
ax.set_xlabel("Truck-vs-non-truck fatality multiplier (identical data)")
ax.set_title("Same data, two specifications: a manufactured ~3.9x multiplier vs the honest ~1.9x ratio",
             fontsize=10.2,weight="bold")
for s in ("top","right","left"): ax.spines[s].set_visible(False)
ax.tick_params(left=False)
fig.tight_layout(); fig.savefig(f"{OUT}/fig_r3_wrong_vs_right.png",dpi=200,bbox_inches="tight")
print("\nSaved: r3_wrong_vs_right.csv, fig_r3_wrong_vs_right.png")
