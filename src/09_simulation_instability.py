"""V2b Experiment A: instability surface of the count-on-count multiplier.
Grid (rare-event rate r_B x denominator smoothness lambda_B), two truth panels
(tau=1.0, 1.9). Metric: P(spurious multiplier >= 1.5 x truth) + 95% width + sign-flip.
Vectorized via Frisch-Waugh (FE annihilator precomputed); numpy/openpyxl/matplotlib."""
import numpy as np, pandas as pd, os
from numpy.linalg import pinv
from openpyxl import load_workbook
import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
OUT="outputs"; RNG=np.random.default_rng(7); R=6000
def load(fn):
    ws=load_workbook(fn,data_only=False).active; rows=[]
    for r in ws.iter_rows(min_row=2,values_only=True):
        yr,tod,cty,dow,typ,tot=r[:6]
        rows.append((str(yr),(tod or"").strip(),(cty or"").strip(),(dow or"").strip(),(typ or"").strip(),0.0 if tot is None else float(tot)))
    return pd.DataFrame(rows,columns=["Year","TOD","County","DOW","Type","Total"])
def cy(df,a,b):
    L=df[(df.TOD=="")&(df.DOW!="")&(df.County!="")&(df.Type.isin(["Fatal","Personal Injury","Property Damage"]))]
    return pd.concat([L.groupby(["County","Year"]).Total.sum().rename(a),
                      L[L.Type=="Fatal"].groupby(["County","Year"]).Total.sum().rename(b)],axis=1).fillna(0).reset_index()
T=cy(load("data/05_0506_Day_of_Week_Time_of_Day.xlsx"),"tcr","tf")
A=cy(load("data/01_0108_Day_of_Week_Time_of_Day.xlsx"),"acr","af")
m=T.merge(A,on=["County","Year"]); m["ntcr"]=m.acr-m.tcr
base_t=m.groupby("County").tcr.mean().values; base_nt=m.groupby("County").ntcr.mean().values
G=len(base_t); Tn=6; N=G*Tn; cid=np.repeat(np.arange(G),Tn); yr=np.tile(np.arange(Tn),G)
# FE annihilator M = I - Z(Z'Z)^-1 Z'
Z=np.column_stack([np.ones(N)]+[(cid==c).astype(float) for c in range(1,G)]+[(yr==t).astype(float) for t in range(1,Tn)])
M=np.eye(N)-Z@pinv(Z)
lamA=0.153
def fe_slope_vec(C,F):           # C,F: (R,N); returns (R,) FE slopes via FWL
    Cx=C@M.T                     # M symmetric => C@M
    return np.sum(Cx*F,axis=1)/np.sum(Cx*Cx,axis=1)
def cell(rB,lamB,tau):
    Ct=RNG.poisson(base_t[cid]*np.exp(lamA*RNG.standard_normal((R,N)))).astype(float)
    Cn=RNG.poisson(base_nt[cid]*np.exp(lamB*RNG.standard_normal((R,N)))).astype(float)
    Ct=np.maximum(Ct,1); Cn=np.maximum(Cn,1)
    Ft=RNG.poisson(rB*tau*Ct).astype(float); Fn=RNG.poisson(rB*Cn).astype(float)
    wrong=fe_slope_vec(Ct,Ft)/fe_slope_vec(Cn,Fn)
    right=(Ft.sum(1)/Ct.sum(1))/(Fn.sum(1)/Cn.sum(1))
    return wrong,right
rBs=[0.5e-3,1e-3,2.2e-3,4e-3,8e-3]; lamBs=[0.04,0.07,0.10,0.15,0.22]
rows=[]; surf={1.0:np.zeros((len(lamBs),len(rBs))),1.9:np.zeros((len(lamBs),len(rBs)))}
for tau in (1.0,1.9):
    for i,lb in enumerate(lamBs):
        for j,rb in enumerate(rBs):
            w,r=cell(rb,lb,tau); p=np.mean(w>=1.5*tau)
            surf[tau][i,j]=p
            rows.append(dict(tau=tau,rB_per1k=rb*1000,lamB=lb,p_ge_1p5x=round(p,3),
                median_wrong=round(np.nanmedian(w),2),width95=round(np.nanpercentile(w,97.5)-np.nanpercentile(w,2.5),2),
                p_signflip=round(np.mean(w<0),3),median_right=round(np.nanmedian(r),2),
                width95_right=round(np.nanpercentile(r,97.5)-np.nanpercentile(r,2.5),2)))
df=pd.DataFrame(rows); df.to_csv(f"{OUT}/v2_instability_grid.csv",index=False)
# NYMTC cell + observed location
ny=cell(2.2e-3,0.10,1.9); pobs=np.mean(ny[0]>=3.84)
print("NYMTC cell (rB=2.2/1k, lamB=0.10, tau=1.9): P(>=1.5x)=%.2f  P(>=3.84 observed)=%.3f  median_wrong=%.2f  median_right=%.2f"%(
    np.mean(ny[0]>=1.5*1.9),pobs,np.nanmedian(ny[0]),np.nanmedian(ny[1])))
print(df.to_string(index=False))
# figure: two heatmaps
import matplotlib.patheffects as _pe
fig,axs=plt.subplots(2,1,figsize=(8.0,7.2))
for ax,tau,ttl in zip(axs,(1.0,1.9),("\u03C4 = 1.0  (no true difference)","\u03C4 = 1.9  (NYMTC truth)")):
    im=ax.imshow(surf[tau],origin="lower",aspect="auto",cmap="RdYlBu_r",vmin=0,vmax=0.4,
                 extent=[-0.5,len(rBs)-0.5,-0.5,len(lamBs)-0.5])
    ax.set_xticks(range(len(rBs))); ax.set_xticklabels([f"{x*1000:g}" for x in rBs])
    ax.set_yticks(range(len(lamBs))); ax.set_yticklabels([f"{y:.2f}" for y in lamBs])
    ax.set_xlabel("rare-event rate (fatal per 1,000)"); ax.set_ylabel("denominator smoothness (CV)")
    ax.set_title(ttl,fontsize=11,weight="bold"); ax.tick_params(labelsize=9)
    for i in range(len(lamBs)):
        for j in range(len(rBs)):
            _t=ax.text(j,i,f"{surf[tau][i,j]:.2f}",ha="center",va="center",fontsize=11,color="black",zorder=5)
            _t.set_path_effects([_pe.withStroke(linewidth=2.2,foreground="white")])
    # NYMTC marker: rB=2.2 (idx2), lamB=0.10 (idx2)
    # NYMTC regime marker in the upper-right of its cell (idx 2,2) to avoid the centered value
    ax.plot(2.30,2.30,marker="*",ms=12,color="black",mec="white",mew=0.9,zorder=6)
    fig.colorbar(im,ax=ax,fraction=0.046,pad=0.04,label="P(multiplier \u2265 1.5\u00D7 truth)")
axs[0].set_xlabel("")
fig.suptitle("Instability surface: how often the count-on-count specification manufactures a \u22651.5\u00D7 multiplier\n"
             "(\u2605 = NYMTC regime; Poisson/worst-case dispersion). The estimator is ~unbiased in median but unidentified.",
             fontsize=10.5,weight="bold")
fig.tight_layout(rect=[0,0,1,0.93]); fig.savefig(f"{OUT}/fig_v2_instability_surface.png",dpi=300,bbox_inches="tight")
print("\nSaved v2_instability_grid.csv, fig_v2_instability_surface.png")
