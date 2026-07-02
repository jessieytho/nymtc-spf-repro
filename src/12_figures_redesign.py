"""Rebuild Fig 2 hybrid with the numeric table in a dedicated panel below the plot,
so the x-axis is unobstructed."""
import pandas as pd, numpy as np
from numpy.linalg import pinv
from openpyxl import load_workbook
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import gridspec
OUT="outputs"; RNG=np.random.default_rng(7)
def load(fn):
    ws=load_workbook(fn,data_only=False).active; rows=[]
    for r in ws.iter_rows(min_row=2,values_only=True):
        yr,tod,cty,dow,typ,tot=r[:6]
        rows.append((str(yr),(tod or"").strip(),(cty or"").strip(),(dow or"").strip(),(typ or"").strip(),0.0 if tot is None else float(tot)))
    return pd.DataFrame(rows,columns=["Year","TOD","County","DOW","Type","Total"])
def cyf(df,a,b):
    L=df[(df.TOD=="")&(df.DOW!="")&(df.County!="")&(df.Type.isin(["Fatal","Personal Injury","Property Damage"]))]
    return pd.concat([L.groupby(["County","Year"]).Total.sum().rename(a),
                      L[L.Type=="Fatal"].groupby(["County","Year"]).Total.sum().rename(b)],axis=1).fillna(0).reset_index()
T=cyf(load("data/05_0506_Day_of_Week_Time_of_Day.xlsx"),"tcr","tf")
A=cyf(load("data/01_0108_Day_of_Week_Time_of_Day.xlsx"),"acr","af")
m=T.merge(A,on=["County","Year"]); m["ntcr"]=m.acr-m.tcr
base_t=m.groupby("County").tcr.mean().values; base_nt=m.groupby("County").ntcr.mean().values
G=len(base_t); Tn=6; N=G*Tn; cid=np.repeat(np.arange(G),Tn); yr=np.tile(np.arange(Tn),G)
Z=np.column_stack([np.ones(N)]+[(cid==c).astype(float) for c in range(1,G)]+[(yr==t).astype(float) for t in range(1,Tn)])
M=np.eye(N)-Z@pinv(Z)
rB=2.2e-3; tau=1.91; lamA=0.153; lamB=0.10; Rn=6000
Ct=np.maximum(RNG.poisson(base_t[cid]*np.exp(lamA*RNG.standard_normal((Rn,N)))),1).astype(float)
Cn=np.maximum(RNG.poisson(base_nt[cid]*np.exp(lamB*RNG.standard_normal((Rn,N)))),1).astype(float)
Ft=RNG.poisson(rB*tau*Ct).astype(float); Fn=RNG.poisson(rB*Cn).astype(float)
w=(np.sum((Ct@M)*Ft,1)/np.sum((Ct@M)**2,1))/(np.sum((Cn@M)*Fn,1)/np.sum((Cn@M)**2,1))

fig=plt.figure(figsize=(7.2,5.1))
gs=gridspec.GridSpec(2,1,height_ratios=[3.3,1.05],hspace=0.66)
ax=fig.add_subplot(gs[0])
wc=np.clip(w,-2,8)
vp=ax.violinplot([wc],positions=[1],vert=False,widths=0.80,showextrema=False)
for b in vp['bodies']: b.set_facecolor("#c0392b"); b.set_alpha(0.30)
ax.plot(3.84,1,marker="D",ms=12,color="#7b1010",mec="white",mew=1.3,zorder=5)
ax.annotate("observed 3.84\u00D7\n(\u224893rd pct of noise)",(3.84,1),xytext=(4.25,1.36),fontsize=10.5,color="#7b1010",
            arrowprops=dict(arrowstyle="->",color="#7b1010",lw=1.1))
ax.errorbar(1.91,0,xerr=[[1.91-1.44],[2.57-1.91]],fmt="o",ms=12,color="#1f4e79",capsize=6,lw=2.6,zorder=5)
ax.axvline(1.91,ls="--",color="#2f855a",lw=1.7); ax.text(1.91,1.78,"true conditional ratio 1.91\u00D7",
            ha="center",fontsize=10.5,color="#2f855a")
ax.axvline(1.0,ls=":",color="#888",lw=1.2)
ax.set_yticks([0,1]); ax.set_yticklabels(["Conditional\nfatal-per-crash\n(right)","Count-on-count\nFE slope ratio\n(wrong)"],fontsize=11)
ax.set_ylim(-0.7,2.10); ax.set_xlim(-2,7)
ax.tick_params(axis="x",labelsize=11.5)
ax.set_xlabel("Truck-vs-non-truck lethality multiplier",fontsize=12.5)
ax.set_title("Same data, two specifications: a tight valid interval\nvs. an unidentified smear of noise",
             fontsize=12,weight="bold")
for s in ("top","right"): ax.spines[s].set_visible(False)
# table panel below
axt=fig.add_subplot(gs[1]); axt.axis("off")
tab=axt.table(cellText=[["1.91\u00D7","[1.44, 2.57]  \u2014 valid","honest severity contrast"],
                        ["3.84\u00D7","none \u2014 unidentified","noise draw (\u224893rd pct)"]],
              rowLabels=["Right: conditional ","Wrong: count-on-count "],
              colLabels=["Estimate","95% interval","Verdict"],
              cellLoc="center",colLoc="center",loc="center")
tab.auto_set_font_size(False); tab.set_fontsize(10.5); tab.scale(1,1.65)
for (r,c),cell in tab.get_celld().items():
    cell.set_edgecolor("#cccccc")
    if r==0: cell.set_facecolor("#e8eef5"); cell.set_text_props(weight="bold")
fig.savefig(f"{OUT}/fig_v7_fig2_hybrid.png",dpi=300,bbox_inches="tight")
print("Fig2 hybrid saved")
